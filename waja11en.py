import re
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
INPUT_FILE = "lotwreport.adi"
OUTPUT_MATRIX_CSV = "waja_lotw_report.csv"
OUTPUT_MATRIX_EXCEL = "waja_lotw_report.xlsx"
OUTPUT_ALL_QSOs_CSV = "all_japan_qsos.csv"
OUTPUT_MISSING_EXCEL = "missing_prefectures.xlsx"
OUTPUT_APPLICATION_TXT = "waja_jarl_application.txt"
OUTPUT_MISSING_BASIC_TXT = "missing_basic_waja.txt"
OUTPUT_MISSING_BANDS_TXT = "missing_by_band.txt"
OUTPUT_TOP_STATIONS_TXT = "top_active_stations.txt"

# Official user list (01-47) - 100% CORRECT
PREFECTURES = {
    "01": "Hokkaido", "02": "Aomori", "03": "Iwate", "04": "Akita", "05": "Yamagata",
    "06": "Miyagi", "07": "Fukushima", "08": "Niigata", "09": "Nagano", "10": "Tokyo",
    "11": "Kanagawa", "12": "Chiba", "13": "Saitama", "14": "Ibaraki", "15": "Tochigi",
    "16": "Gunma", "17": "Yamanashi", "18": "Shizuoka", "19": "Gifu", "20": "Aichi",
    "21": "Mie", "22": "Kyoto", "23": "Shiga", "24": "Nara", "25": "Osaka",
    "26": "Wakayama", "27": "Hyogo", "28": "Toyama", "29": "Fukui", "30": "Ishikawa",
    "31": "Okayama", "32": "Shimane", "33": "Yamaguchi", "34": "Tottori", "35": "Hiroshima",
    "36": "Kagawa", "37": "Tokushima", "38": "Ehime", "39": "Kochi", "40": "Fukuoka",
    "41": "Saga", "42": "Nagasaki", "43": "Kumamoto", "44": "Oita", "45": "Miyazaki",
    "46": "Kagoshima", "47": "Okinawa"
}

# Bands to track from 160m to 6m
BANDS_LIST = ["160M", "80M", "40M", "30M", "20M", "17M", "15M", "12M", "10M", "6M"]

def stream_adif_records(file_path):
    """Efficiently reads large ADIF files line by line."""
    current_record = []
    try:
        with open(file_path, "r", encoding="latin-1") as f:
            for line in f:
                current_record.append(line)
                if "<eor>" in line.lower():
                    yield "".join(current_record)
                    current_record = []
    except FileNotFoundError:
        print(f"[-] Error: File '{file_path}' not found in this directory!")
        return

def parse_adif_field(record, field_name):
    """Extracts the value of a specific ADIF tag."""
    pattern = rf"<{field_name}:?\d*>[^<\s]+"
    match = re.search(pattern, record, re.IGNORECASE)
    if match:
        full_tag = match.group(0)
        value = full_tag.split(">")[-1].strip()
        return value
    return None

def extract_prefecture(record):
    """Extracts the 2-digit JARL WAJA prefecture code from STATE or CNTY fields."""
    state = parse_adif_field(record, "STATE")
    if state:
        match = re.match(r"^(\d{1,2})", state)
        if match:
            return f"{int(match.group(1)):02d}"
            
    cnty = parse_adif_field(record, "CNTY")
    if cnty:
        match = re.match(r"^(\d{1,2})", cnty)
        if match:
            return f"{int(match.group(1)):02d}"
            
    return None

def apply_excel_formatting(file_path, sheet_name, criteria_missing, criteria_worked):
    """Applies beautiful light red and light green fills to Excel cells based on data."""
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_font = Font(color="9C0006", bold=True)
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_font = Font(color="006100")
            
            max_row = worksheet.max_row
            cell_range = f"C2:L{max_row}"
            
            worksheet.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=[f'"{criteria_missing}"'], fill=red_fill, font=red_font))
            worksheet.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=[f'"{criteria_worked}"'], fill=green_fill, font=green_font))
            
            for col_idx, col in enumerate(worksheet.columns, start=1):
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
    except Exception as e:
        print(f"[!] Warning: Could not style Excel file automatically: {e}")

def main():
    print(f"[*] Starting scan of file '{INPUT_FILE}'...")
    
    matrix_data = {
        code: {b: "." for b in BANDS_LIST} 
        for code in PREFECTURES.keys()
    }
    
    all_qsos_list = []
    total_qsos = 0
    japan_qsos = 0
    
    earliest_qsos = {}
    
    for record in stream_adif_records(INPUT_FILE):
        total_qsos += 1
        
        if total_qsos % 25000 == 0:
            print(f"[i] Read {total_qsos} records...")

        dxcc = parse_adif_field(record, "DXCC")
        
        if dxcc == "339":
            band = parse_adif_field(record, "BAND")
            if band:
                band = band.upper()
                if band in BANDS_LIST:
                    pref_code = extract_prefecture(record)
                    
                    if pref_code in PREFECTURES:
                        callsign = parse_adif_field(record, "CALL")
                        if callsign:
                            callsign = callsign.upper()
                            mode = parse_adif_field(record, "MODE") or "UNKNOWN"
                            qso_date = parse_adif_field(record, "QSO_DATE") or ""
                            time_on = parse_adif_field(record, "TIME_ON") or ""
                            
                            formatted_date = f"{qso_date[:4]}-{qso_date[4:6]}-{qso_date[6:8]}" if len(qso_date) == 8 else qso_date
                            formatted_time = f"{time_on[:2]}:{time_on[2:4]}:{time_on[4:6]}" if len(time_on) == 6 else time_on
                            
                            current_entry = matrix_data[pref_code][band]
                            if current_entry == ".":
                                matrix_data[pref_code][band] = callsign
                            elif callsign not in current_entry:
                                matrix_data[pref_code][band] += f", {callsign}"
                            
                            all_qsos_list.append({
                                "Callsign": callsign,
                                "JARL_Code": pref_code,
                                "Prefecture": PREFECTURES[pref_code],
                                "Band": band,
                                "Mode": mode.upper(),
                                "Date": formatted_date,
                                "Time": formatted_time
                            })
                            
                            qso_timestamp = qso_date + time_on
                            qso_info = {
                                "Callsign": callsign,
                                "Band": band,
                                "Mode": mode.upper(),
                                "Date": formatted_date,
                                "Time": formatted_time,
                                "timestamp": qso_timestamp
                            }
                            
                            if pref_code not in earliest_qsos:
                                earliest_qsos[pref_code] = qso_info
                            else:
                                if qso_timestamp < earliest_qsos[pref_code]["timestamp"]:
                                    earliest_qsos[pref_code] = qso_info
                                    
                            japan_qsos += 1

    print(f"\n[+] Analysis complete! Total ADIF records checked: {total_qsos}")
    print(f"[+] Found {japan_qsos} confirmed Japan QSOs with a valid JARL prefecture.")

    # --- 1. GENERATE MATRIX REPORTS ---
    rows_matrix = []
    for code, pref_name in sorted(PREFECTURES.items()):
        row = {"Code": code, "Prefecture": pref_name}
        row.update(matrix_data[code])
        rows_matrix.append(row)
        
    df_matrix = pd.DataFrame(rows_matrix)
    df_matrix.to_csv(OUTPUT_MATRIX_CSV, index=False, encoding="utf-8-sig")
    df_matrix.to_excel(OUTPUT_MATRIX_EXCEL, index=False, sheet_name="WAJA Report")
    apply_excel_formatting(OUTPUT_MATRIX_EXCEL, "WAJA Report", criteria_missing=".", criteria_worked="") 

    # --- 2. GENERATE MISSING PREFECTURES EXCEL REPORT ---
    rows_missing = []
    band_stats = {b: 0 for b in BANDS_LIST}
    missing_by_band_dict = {b: [] for b in BANDS_LIST}
    
    any_band_worked_prefectures = set()
    five_band_prefectures = []
    nine_band_prefectures = []
    
    for code, pref_name in sorted(PREFECTURES.items()):
        row_m = {"Code": code, "Prefecture": pref_name}
        worked_bands_count = 0
        
        for band in BANDS_LIST:
            if matrix_data[code][band] == ".":
                row_m[band] = "MISSING"
                missing_by_band_dict[band].append((code, pref_name))
            else:
                row_m[band] = "WORKED"
                band_stats[band] += 1
                worked_bands_count += 1
                any_band_worked_prefectures.add(code)
                
        if worked_bands_count >= 5:
            five_band_prefectures.append(f"{code} ({pref_name})")
        if worked_bands_count >= 9:
            nine_band_prefectures.append(f"{code} ({pref_name})")
            
        rows_missing.append(row_m)
        
    df_missing = pd.DataFrame(rows_missing)
    df_missing.to_excel(OUTPUT_MISSING_EXCEL, index=False, sheet_name="Missing Prefectures")
    apply_excel_formatting(OUTPUT_MISSING_EXCEL, "Missing Prefectures", criteria_missing="MISSING", criteria_worked="WORKED")

    # --- 3. SAVE ALL INDIVIDUAL QSOs LIST ---
    if all_qsos_list:
        df_all_qsos = pd.DataFrame(all_qsos_list)
        df_all_qsos = df_all_qsos.sort_values(by=["JARL_Code", "Band", "Date"])
        df_all_qsos.to_csv(OUTPUT_ALL_QSOs_CSV, index=False, encoding="utf-8-sig")

    # --- 4. GENERATE JARL WAJA APPLICATION FORM (TXT) ---
    try:
        with open(OUTPUT_APPLICATION_TXT, "w", encoding="utf-8") as app_file:
            app_file.write("======================================================\n")
            app_file.write("           JARL WAJA AWARD APPLICATION FORM           \n")
            app_file.write("======================================================\n\n")
            app_file.write("Applicant Callsign: 9A5CW\n")
            app_file.write(f"Total Prefectures Confirmed: {len(earliest_qsos)} / 47\n")
            app_file.write("Verification Source: ARRL LoTW (Logbook of The World)\n\n")
            app_file.write(f"{'No.':<4} {'Prefecture':<15} {'Callsign':<10} {'Date':<12} {'Time':<8} {'Band':<6} {'Mode':<6}\n")
            app_file.write("-" * 65 + "\n")
            
            for code, pref_name in sorted(PREFECTURES.items()):
                if code in earliest_qsos:
                    q = earliest_qsos[code]
                    app_file.write(f"{code:<4} {pref_name:<15} {q['Callsign']:<10} {q['Date']:<12} {q['Time']:<8} {q['Band']:<6} {q['Mode']:<6}\n")
                else:
                    app_file.write(f"{code:<4} {pref_name:<15} {'[ MISSING - NOT WORKED YET ]':<45}\n")
                    
            app_file.write("\n======================================================\n")
            app_file.write("Generated automatically via Python ADIF WAJA Parser.\n")
        print(f"[+] Službena JARL prijavna lista stvorena u: '{OUTPUT_APPLICATION_TXT}'")
    except Exception as e:
        print(f"[!] Warning: Could not generate text application file: {e}")

    # --- 5. SAVE MISSING FOR BASIC WAJA TO SEPARATE TXT ---
    missing_all_bands = [(code, name) for code, name in sorted(PREFECTURES.items()) if code not in any_band_worked_prefectures]
    try:
        with open(OUTPUT_MISSING_BASIC_TXT, "w", encoding="utf-8") as f:
            f.write("======================================================\n")
            f.write("         MISSING PREFECTURES FOR BASIC WAJA           \n")
            f.write("======================================================\n\n")
            f.write(f"Total Missing for Basic WAJA: {len(missing_all_bands)} / {len(PREFECTURES)}\n\n")
            f.write(f"{'Code':<6} {'Prefecture Name':<20}\n")
            f.write("-" * 30 + "\n")
            if missing_all_bands:
                for code, name in missing_all_bands:
                    f.write(f"{code:<6} {name:<20}\n")
            else:
                f.write("[CONGRATULATIONS] All 47 prefectures worked at least once!\n")
            f.write("\n======================================================\n")
        print(f"[+] Missing basic WAJA list saved to: '{OUTPUT_MISSING_BASIC_TXT}'")
    except Exception as e:
        print(f"[!] Warning: Could not save missing basic TXT: {e}")

    # --- 6. SAVE MISSING PREFECTURES BY BAND TO SEPARATE TXT ---
    try:
        with open(OUTPUT_MISSING_BANDS_TXT, "w", encoding="utf-8") as f:
            f.write("======================================================\n")
            f.write("         MISSING PREFECTURES BY BAND REPORT           \n")
            f.write("======================================================\n\n")
            for band in BANDS_LIST:
                missing_list = missing_by_band_dict[band]
                worked_count = band_stats[band]
                pct = (worked_count / len(PREFECTURES)) * 100
                f.write(f"--- BAND: {band} (Worked: {worked_count}/47, Missing: {len(missing_list)}/47, {pct:.1f}%) ---\n")
                if missing_list:
                    for code, name in missing_list:
                        f.write(f"  [{code}] {name}\n")
                else:
                    f.write("  [CONGRATULATIONS] All 47 prefectures worked on this band!\n")
                f.write("\n")
            f.write("======================================================\n")
        print(f"[+] Missing by band report saved to: '{OUTPUT_MISSING_BANDS_TXT}'")
    except Exception as e:
        print(f"[!] Warning: Could not save missing by band TXT: {e}")

# --- 7. GENERATE TOP ACTIVE STATIONS REPORT ---
    try:
        station_stats = {}
        for qso in all_qsos_list:
            call = qso['Callsign']
            pref = qso['JARL_Code']
            band = qso['Band']
            if call not in station_stats:
                station_stats[call] = {'total_qsos': 0, 'bands': set(), 'prefectures': set()}
            station_stats[call]['total_qsos'] += 1
            station_stats[call]['bands'].add(band)
            station_stats[call]['prefectures'].add(pref)

        # Sortiranje: po broju opsega, pa broju prefektura, pa ukupnom broju QSO-a
        sorted_stations = sorted(
            station_stats.items(),
            key=lambda x: (len(x[1]['bands']), len(x[1]['prefectures']), x[1]['total_qsos']),
            reverse=True
        )

        with open(OUTPUT_TOP_STATIONS_TXT, "w", encoding="utf-8") as f:
            f.write("========================================================================================\n")
            f.write("                       TOP ACTIVE JAPANESE STATIONS ACROSS BANDS                        \n")
            f.write("========================================================================================\n\n")
            f.write(f"{'Callsign':<10} {'QSOs':<6} {'Prefs Worked':<14} {'Bands Worked'}\n")
            f.write("-" * 88 + "\n")
            
            for call, stats in sorted_stations[:100]:  # Prikaz top 100 stanica
                # Sortiranje opsega logičnim redoslijedom ako je moguće, ili abecedno
                sorted_bands = sorted(list(stats['bands']))
                bands_str = ", ".join(sorted_bands)
                
                sorted_prefs = sorted(list(stats['prefectures']))
                prefs_str = ", ".join(sorted_prefs)
                
                f.write(f"{call:<10} {stats['total_qsos']:<6} {len(stats['prefectures']):<14} {bands_str}\n")
                f.write(f"           Prefs: {prefs_str}\n")
                f.write("-" * 88 + "\n")
                
            f.write("========================================================================================\n")
        print(f"[+] Top active stations report saved to: '{OUTPUT_TOP_STATIONS_TXT}'")
    except Exception as e:
        print(f"[!] Warning: Could not generate top active stations report: {e}")

# --- 8. GENERATE MISSING BAND LEADS REPORT ---
    try:
        station_to_prefs = {}
        station_to_bands = {}
        for qso in all_qsos_list:
            call = qso['Callsign']
            pref = qso['JARL_Code']
            band = qso['Band']
            
            if call not in station_to_prefs:
                station_to_prefs[call] = set()
            station_to_prefs[call].add(pref)
            
            if call not in station_to_bands:
                station_to_bands[call] = set()
            station_to_bands[call].add(band)

        pref_to_stations = {}
        for call, prefs in station_to_prefs.items():
            for p in prefs:
                if p not in pref_to_stations:
                    pref_to_stations[p] = set()
                pref_to_stations[p].add(call)

        OUTPUT_LEADS_TXT = "missing_band_leads.txt"
        with open(OUTPUT_LEADS_TXT, "w", encoding="utf-8") as f:
            f.write("========================================================================================\n")
            f.write("                     POTENCIJALNE STANICE ZA NEDOSTAJUĆE PREFEKTURE PO BANDOVRIBUTO        \n")
            f.write("========================================================================================\n")
            f.write("Ovaj izvještaj pokazuje stanice koje ste već radili na drugim bandovima, a dolaze iz\n")
            f.write("prefektura koje su vam možda 'rupa' na nekom specifičnom bandu.\n")
            f.write("========================================================================================\n\n")

            confirmed_band_pref = {b: set() for b in BANDS_LIST}
            for qso in all_qsos_list:
                b = qso['Band']
                p = qso['JARL_Code']
                if b in confirmed_band_pref:
                    confirmed_band_pref[b].add(p)

            all_prefs = {f"{i:02d}" for i in range(1, 48)}

            for band in BANDS_LIST:
                missing_prefs_for_band = sorted(list(all_prefs - confirmed_band_pref[band]))
                if not missing_prefs_for_band:
                    continue
                
                f.write(f"--- BAND: {band} (Nedostaje {len(missing_prefs_for_band)} prefektura) ---\n")
                
                found_leads_for_band = False
                for pref in missing_prefs_for_band:
                    candidates = pref_to_stations.get(pref, set())
                    active_candidates = []
                    for c in candidates:
                        c_bands = station_to_bands[c]
                        if band not in c_bands:
                            active_candidates.append((c, sorted(list(c_bands))))
                    
                    if active_candidates:
                        found_leads_for_band = True  # Ovdje je bila greška (malo true -> veliko True)
                        f.write(f"  Prefektura {pref} ({PREFECTURES.get(pref, '')}):\n")
                        for call, cbands in active_candidates:
                            bands_str = ", ".join(cbands)
                            f.write(f"    -> Stanica {call} (već radili na: {bands_str})\n")
                
                if not found_leads_for_band:
                    f.write("  (Nema pronađenih poznatih stanica iz nedostajućih prefektura na drugim bandovima)\n")
                f.write("\n")

            f.write("========================================================================================\n")
        print(f"[+] Missing band leads report saved to: '{OUTPUT_LEADS_TXT}'")
    except Exception as e:
        print(f"[!] Warning: Could not generate leads report: {e}")
                        
    # --- LIVE SUMMARY STATISTICS IN TERMINAL ---
    print("\n================ WAJA BAND STATISTICS ================")
    total_prefectures = len(PREFECTURES)
    for band in BANDS_LIST:
        worked = band_stats[band]
        percentage = (worked / total_prefectures) * 100
        print(f"-> Band {band:5}: Confirmed {worked:2d}/{total_prefectures} Prefectures ({percentage:.1f}%)")

    print("\n================ AWARD TRACKER SUMMARY ================")
    print(f"-> Basic WAJA Status  : Worked {len(any_band_worked_prefectures)}/{total_prefectures} unique Prefectures.")
    print(f"-> 5-Band WAJA Status : {len(five_band_prefectures)} Prefectures completed on 5+ bands.")
    if five_band_prefectures:
        print(f"   Completed: {', '.join(five_band_prefectures[:5])} ...")
    print(f"-> 9-Band WAJA Status : {len(nine_band_prefectures)} Prefectures completed on 9+ bands.")
    if nine_band_prefectures:
        print(f"   Completed: {', '.join(nine_band_prefectures)}")

    if missing_all_bands:
        print(f"-> Missing for basic WAJA ({len(missing_all_bands)} total) saved to '{OUTPUT_MISSING_BASIC_TXT}'")
    else:
        print("-> Congratulations! You have worked all 47 prefectures at least once!")
    print(f"-> Missing by band breakdown saved to '{OUTPUT_MISSING_BANDS_TXT}'")
    print(f"-> Top active stations report saved to '{OUTPUT_TOP_STATIONS_TXT}'")
    print("======================================================")

if __name__ == "__main__":
    main()